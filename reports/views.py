from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from datetime import datetime, timedelta, date
from receipts.models import Receipt
from expenses.models import Expense
from students.models import Student

@login_required
def reports_dashboard(request):
    """Main reports dashboard"""
    return render(request, 'reports/dashboard.html')

@login_required
def daily_cash_report(request):
    """Daily cash position report"""
    report_date = request.GET.get('date', date.today())
    if isinstance(report_date, str):
        report_date = datetime.strptime(report_date, '%Y-%m-%d').date()
    
    receipts = Receipt.objects.filter(payment_date=report_date, is_voided=False)
    total_collected_lrd = receipts.aggregate(Sum('amount_lrd'))['amount_lrd__sum'] or 0
    total_collected_usd = receipts.aggregate(Sum('amount_usd'))['amount_usd__sum'] or 0
    
    expenses = Expense.objects.filter(expense_date=report_date)
    total_expenses_lrd = expenses.aggregate(Sum('amount_lrd'))['amount_lrd__sum'] or 0
    total_expenses_usd = expenses.aggregate(Sum('amount_usd'))['amount_usd__sum'] or 0
    
    exchange_rate = 200
    closing_balance = (total_collected_lrd - total_expenses_lrd) + \
                      ((total_collected_usd - total_expenses_usd) * exchange_rate)
    
    context = {
        'report_date': report_date,
        'total_collected_lrd': total_collected_lrd,
        'total_collected_usd': total_collected_usd,
        'total_expenses_lrd': total_expenses_lrd,
        'total_expenses_usd': total_expenses_usd,
        'closing_balance': closing_balance,
        'receipts': receipts,
        'expenses': expenses,
    }
    return render(request, 'reports/daily_cash.html', context)

@login_required
def weekly_collections(request):
    """Weekly fee collection by class"""
    from datetime import date, timedelta
    
    week_start = request.GET.get('week_start', date.today() - timedelta(days=7))
    if isinstance(week_start, str):
        week_start = datetime.strptime(week_start, '%Y-%m-%d').date()
    week_end = week_start + timedelta(days=6)
    
    receipts = Receipt.objects.filter(
        payment_date__gte=week_start,
        payment_date__lte=week_end,
        is_voided=False
    )
    
    # Group by student's class - FIXED: use get_class_name() method
    collections_by_class = {}
    for receipt in receipts:
        # Use the method instead of property
        class_name = receipt.student.get_class_name()
        if class_name not in collections_by_class:
            collections_by_class[class_name] = {'lrd': 0, 'usd': 0, 'count': 0}
        collections_by_class[class_name]['lrd'] += float(receipt.amount_lrd)
        collections_by_class[class_name]['usd'] += float(receipt.amount_usd)
        collections_by_class[class_name]['count'] += 1
    
    context = {
        'week_start': week_start,
        'week_end': week_end,
        'collections_by_class': collections_by_class,
        'total_lrd': sum(c['lrd'] for c in collections_by_class.values()),
        'total_usd': sum(c['usd'] for c in collections_by_class.values()),
    }
    return render(request, 'reports/weekly_collections.html', context)

@login_required
def termly_summary(request):
    """Termly income/expenditure summary"""
    term = request.GET.get('term', 'Term 1')
    year = request.GET.get('year', date.today().year)
    
    current_month = date.today().month
    month_start = date(date.today().year, current_month, 1)
    month_end = date.today()
    
    receipts = Receipt.objects.filter(
        payment_date__gte=month_start,
        payment_date__lte=month_end,
        is_voided=False
    )
    expenses = Expense.objects.filter(
        expense_date__gte=month_start,
        expense_date__lte=month_end
    )
    
    total_income_lrd = receipts.aggregate(Sum('amount_lrd'))['amount_lrd__sum'] or 0
    total_income_usd = receipts.aggregate(Sum('amount_usd'))['amount_usd__sum'] or 0
    total_expenses_lrd = expenses.aggregate(Sum('amount_lrd'))['amount_lrd__sum'] or 0
    total_expenses_usd = expenses.aggregate(Sum('amount_usd'))['amount_usd__sum'] or 0
    
    context = {
        'term': term,
        'year': year,
        'period_start': month_start,
        'period_end': month_end,
        'total_income_lrd': total_income_lrd,
        'total_income_usd': total_income_usd,
        'total_expenses_lrd': total_expenses_lrd,
        'total_expenses_usd': total_expenses_usd,
        'net_lrd': total_income_lrd - total_expenses_lrd,
        'net_usd': total_income_usd - total_expenses_usd,
    }
    return render(request, 'reports/termly_summary.html', context)

@login_required
def arrears_report(request):
    """Students with outstanding balances"""
    from fees.models import StudentFeeLedger
    
    students_with_balance = []
    students = Student.objects.filter(is_active=True)
    
    for student in students:
        ledger = StudentFeeLedger.objects.filter(student=student).first()
        if ledger and ledger.total_balance_lrd > 0:
            students_with_balance.append({
                'admission_number': student.admission_number,
                'name': student.full_name,
                'class_name': student.get_class_name(),  # Use method
                'parent_phone': student.parent_phone,
                'balance_lrd': ledger.total_balance_lrd,
            })
    
    # Sort by balance (highest first)
    students_with_balance.sort(key=lambda x: x['balance_lrd'], reverse=True)
    
    context = {
        'students': students_with_balance,
        'total_students': len(students_with_balance),
    }
    return render(request, 'reports/arrears.html', context)

@login_required
def missing_receipts_report(request):
    """Report showing missing receipt numbers"""
    from receipts.models import Receipt
    from django.db.models import Max, Min
    
    # Get all receipt numbers
    receipts = Receipt.objects.filter(is_voided=False).order_by('receipt_number')
    
    if not receipts.exists():
        context = {'missing_numbers': [], 'message': 'No receipts found'}
        return render(request, 'reports/missing_receipts.html', context)
    
    first_receipt = receipts.first().receipt_number
    last_receipt = receipts.last().receipt_number
    
    # Get all existing numbers
    existing_numbers = set(receipts.values_list('receipt_number', flat=True))
    
    # Find missing numbers
    missing_numbers = []
    for num in range(first_receipt, last_receipt + 1):
        if num not in existing_numbers:
            missing_numbers.append(num)
    
    context = {
        'missing_numbers': missing_numbers,
        'first_receipt': first_receipt,
        'last_receipt': last_receipt,
        'total_receipts': len(existing_numbers),
        'expected_count': last_receipt - first_receipt + 1,
        'missing_count': len(missing_numbers),
    }
    return render(request, 'reports/missing_receipts.html', context)

@login_required
def missing_receipts_report(request):
    """Report showing missing receipt numbers"""
    from receipts.models import Receipt
    
    receipts = Receipt.objects.filter(is_voided=False).order_by('receipt_number')
    
    if not receipts.exists():
        context = {'missing_numbers': [], 'message': 'No receipts found'}
        return render(request, 'reports/missing_receipts.html', context)
    
    first_receipt = receipts.first().receipt_number
    last_receipt = receipts.last().receipt_number
    
    existing_numbers = set(receipts.values_list('receipt_number', flat=True))
    
    missing_numbers = []
    for num in range(first_receipt, last_receipt + 1):
        if num not in existing_numbers:
            missing_numbers.append(num)
    
    context = {
        'missing_numbers': missing_numbers,
        'first_receipt': first_receipt,
        'last_receipt': last_receipt,
        'total_receipts': len(existing_numbers),
        'expected_count': last_receipt - first_receipt + 1,
        'missing_count': len(missing_numbers),
    }
    return render(request, 'reports/missing_receipts.html', context)


@login_required
def export_audit_summary(request):
    """Export one-page summary for annual audit"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from django.db.models import Sum
    from datetime import datetime
    from receipts.models import Receipt
    from expenses.models import Expense
    from students.models import Student
    from django.http import HttpResponse
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit Summary"
    
    ws['A1'] = "SFMS - SCHOOL FINANCIAL MANAGEMENT SYSTEM"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:F1')
    
    ws['A2'] = f"Audit Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws.merge_cells('A2:F2')
    
    ws['A4'] = "SUMMARY STATISTICS"
    ws['A4'].font = Font(bold=True, size=12)
    
    total_students = Student.objects.filter(is_active=True).count()
    total_receipts = Receipt.objects.filter(is_voided=False).count()
    total_collected_lrd = Receipt.objects.filter(is_voided=False).aggregate(Sum('amount_lrd'))['amount_lrd__sum'] or 0
    total_collected_usd = Receipt.objects.filter(is_voided=False).aggregate(Sum('amount_usd'))['amount_usd__sum'] or 0
    total_expenses_lrd = Expense.objects.aggregate(Sum('amount_lrd'))['amount_lrd__sum'] or 0
    total_expenses_usd = Expense.objects.aggregate(Sum('amount_usd'))['amount_usd__sum'] or 0
    
    stats_data = [
        ['Total Active Students', total_students],
        ['Total Receipts Issued', total_receipts],
        ['Total Collections (LRD)', f"{total_collected_lrd:,.2f}"],
        ['Total Collections (USD)', f"{total_collected_usd:,.2f}"],
        ['Total Expenses (LRD)', f"{total_expenses_lrd:,.2f}"],
        ['Total Expenses (USD)', f"{total_expenses_usd:,.2f}"],
        ['Net Surplus (LRD)', f"{total_collected_lrd - total_expenses_lrd:,.2f}"],
    ]
    
    for i, row in enumerate(stats_data, start=6):
        ws[f'A{i}'] = row[0]
        ws[f'B{i}'] = row[1]
        ws[f'A{i}'].font = Font(bold=True)
    
    ws['A15'] = "RECENT RECEIPTS (Last 20)"
    ws['A15'].font = Font(bold=True, size=12)
    
    headers = ['Receipt #', 'Date', 'Student', 'Amount LRD', 'Amount USD', 'Method']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=17, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    recent_receipts = Receipt.objects.filter(is_voided=False).order_by('-receipt_number')[:20]
    row = 18
    for receipt in recent_receipts:
        ws.cell(row=row, column=1, value=receipt.receipt_number)
        ws.cell(row=row, column=2, value=receipt.payment_date.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=3, value=receipt.student.full_name)
        ws.cell(row=row, column=4, value=float(receipt.amount_lrd))
        ws.cell(row=row, column=5, value=float(receipt.amount_usd))
        ws.cell(row=row, column=6, value=receipt.get_payment_method_display())
        row += 1
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 20
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="sfms_audit_summary_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


@login_required
def export_receipts_excel(request):
    """Export all receipts to Excel"""
    import openpyxl
    from openpyxl.styles import Font
    from datetime import datetime
    from receipts.models import Receipt
    from django.http import HttpResponse
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Receipts"
    
    headers = ['Receipt #', 'Date', 'Student', 'Admission #', 'Class', 'Amount LRD', 'Amount USD', 'Method', 'Voided']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
    
    receipts = Receipt.objects.all().order_by('-receipt_number')
    row = 2
    for receipt in receipts:
        ws.cell(row=row, column=1, value=receipt.receipt_number)
        ws.cell(row=row, column=2, value=receipt.payment_date.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=3, value=receipt.student.full_name)
        ws.cell(row=row, column=4, value=receipt.student.admission_number)
        ws.cell(row=row, column=5, value=receipt.student.get_class_name())
        ws.cell(row=row, column=6, value=float(receipt.amount_lrd))
        ws.cell(row=row, column=7, value=float(receipt.amount_usd))
        ws.cell(row=row, column=8, value=receipt.get_payment_method_display())
        ws.cell(row=row, column=9, value='Yes' if receipt.is_voided else 'No')
        row += 1
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="sfms_receipts_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response

@login_required
def send_whatsapp_reminders(request):
    """Generate WhatsApp reminders for parents with overdue fees"""
    from students.models import Student
    from fees.models import StudentFeeLedger
    from core.whatsapp_utils import generate_whatsapp_link, generate_default_reminder, generate_overdue_message
    from core.models import SchoolSetting
    from datetime import date, timedelta
    from django.contrib import messages  # ADD THIS IMPORT
    
    school_name_setting = SchoolSetting.objects.filter(key='school_name').first()
    school_name = school_name_setting.value if school_name_setting else 'SFMS SCHOOL'
    
    defaulters = []
    
    if request.method == 'POST':
        selected_students = request.POST.getlist('selected_students')
        message_type = request.POST.get('message_type', 'reminder')
        
        for student_id in selected_students:
            student = Student.objects.get(id=student_id)
            ledger = StudentFeeLedger.objects.filter(student=student).first()
            
            if ledger and ledger.total_balance_lrd > 0:
                if message_type == 'reminder':
                    message = generate_default_reminder(
                        student.full_name, 
                        ledger.total_balance_lrd,
                        (date.today() + timedelta(days=7)).strftime('%Y-%m-%d'),
                        school_name
                    )
                else:
                    days_overdue = (date.today() - ledger.last_payment_date).days if ledger.last_payment_date else 30
                    message = generate_overdue_message(
                        student.full_name,
                        ledger.total_balance_lrd,
                        days_overdue,
                        school_name
                    )
                
                if student.parent_phone:
                    whatsapp_link = generate_whatsapp_link(student.parent_phone, message)
                    defaulters.append({
                        'student': student,
                        'phone': student.parent_phone,
                        'balance': ledger.total_balance_lrd,
                        'whatsapp_link': whatsapp_link,
                        'message': message
                    })
        
        # Store in session for bulk sending
        request.session['whatsapp_messages'] = [
            {'phone': d['phone'], 'link': d['whatsapp_link'], 'student': d['student'].full_name}
            for d in defaulters
        ]
        
        messages.success(request, f'Prepared {len(defaulters)} WhatsApp messages. Click the links to send.')
        return render(request, 'reports/whatsapp_reminders_result.html', {'defaulters': defaulters})
    
    # GET request - show students with balances
    students = Student.objects.filter(is_active=True)
    for student in students:
        ledger = StudentFeeLedger.objects.filter(student=student).first()
        if ledger and ledger.total_balance_lrd > 0:
            defaulters.append({
                'student': student,
                'balance': ledger.total_balance_lrd,
                'phone': student.parent_phone or 'No phone'
            })
    
    context = {
        'defaulters': defaulters,
        'school_name': school_name,
    }
    return render(request, 'reports/whatsapp_reminders.html', context)
